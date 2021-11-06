import openpyxl


def request_base(phone: str = '', name: str = '', *args) -> list:
    file = 'base.xlsx'
    matches = list()
    wbook = openpyxl.load_workbook(filename=file, read_only=True)
    sheet = wbook.active
    for row in sheet.iter_rows(
        min_row=1, 
        max_row=sheet.max_row, 
        max_col=sheet.max_column
    ):
        get_row = [str(cell.value) for cell in row]
        row_phone = get_row[1]
        row_personal_name = get_row[2]
        if phone and name:
            if phone in row_phone and phone in row_phone:
                matches.append((row_phone, row_personal_name))
            elif phone in row_phone:
                matches.append((row_phone, row_personal_name))
        elif phone and phone in row_phone:
            matches.append((row_phone, row_personal_name))
        elif name and name in row_personal_name:
            matches.append((row_phone, row_personal_name))
    get_records = sheet.max_row
    wbook.close()

    return matches, get_records